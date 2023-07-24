import streamlit as st
import io
import pandas as pd
import os
import sys
import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from emissions_factors import get_factors



#-------------------- STREAMLIT CONTAINERS --------------------

st.title('Reformat ENERGY STAR .xlsx')
st.write('''Upload output .xlsx file from Portfolio Manager and transform the data to work with Carbon Signal. 
         Using **Blend** will attempt to find a complete dataset by merging multiple years of data. 
         Using **Complete** will grab utility data from calendar years with most complete data. Years may vary across utilities.
         Using **Latest** will pull data from the last calendar year (e.g. 2022)''')

form_container = st.container()
spacer1 = st.write('')
spacer2 = st.write('')
button_container = st.empty()
messages = st.empty()

def reset():
    messages.empty()









#-------------------- UTILITY FUNCTIONS --------------------


#Convert energy units, takes in the type (t), converts from unit (a) to unit(b)
def convert(t, a, b):
    #https://portfoliomanager.energystar.gov/pdf/reference/Thermal%20Conversions.pdf
    conversions = {
        'Area': {
            'Sq. M.': {'Sq. Ft.': 10.76}
        },
        'Natural Gas': {
            'ccf (hundred cubic feet)': {'kBtu (thousand Btu)': 102.6},
            'cf (cubic feet)': {'kBtu (thousand Btu)': 1.026},
            'cm (cubic meters)': {'kBtu (thousand Btu)': 36.303},
            'GJ': {'kBtu (thousand Btu)': 947.817},
            'kcf (thousand cubic feet)': {'kBtu (thousand Btu)': 1026},
            'MBtu/ MMBtu/ Dth (million Btu/ dekatherm)': {'kBtu (thousand Btu)': 1000},
            'Mcf (million cibic feet)': {'kBtu (thousand Btu)': 1026000},
            'therms': {'kBtu (thousand Btu)': 100}
        },
        'Propane': {
            'ccf (hundred cubic feet)': {'kBtu (thousand Btu)': 251.6},
            'cf (cubic feet)': {'kBtu (thousand Btu)': 2.516},
            'Gallons (UK)': {'kBtu (thousand Btu)': 110.484},
            'Gallons (US)': {'kBtu (thousand Btu)': 92},
            'GJ': {'kBtu (thousand Btu)': 947.817},
            'kcf (thousand cubic feet)': {'kBtu (thousand Btu)': 2516},
            'Liters': {'kBtu (thousand Btu)': 24.304},
            'MBtu/ MMBtu/ Dth (million Btu/ dekatherm)': {'kBtu (thousand Btu)': 1000}
        },
        'Fuel Oil (No. 2)': {
            'Gallons (UK)': {'kBtu (thousand Btu)': 165.726},
            'Gallons (US)': {'kBtu (thousand Btu)': 138},
            'GJ': {'kBtu (thousand Btu)': 947.817},
            'Liters': {'kBtu (thousand Btu)': 36.456},
            'MBtu/ MMBtu/ Dth (million Btu/ dekatherm)': {'kBtu (thousand Btu)': 1000}
        },
        'Diesel': {
            'Gallons (UK)': {'kBtu (thousand Btu)': 165.726},
            'Gallons (US)': {'kBtu (thousand Btu)': 138},
            'GJ': {'kBtu (thousand Btu)': 947.817},
            'Liters': {'kBtu (thousand Btu)': 36.456},
            'MBtu/ MMBtu/ Dth (million Btu/ dekatherm)': {'kBtu (thousand Btu)': 1000}
        },
        'District Steam': {
            'kg (kilograms)': {'kBtu (thousand Btu)': 2.632},
            'kLbs. (thousand pounds)': {'kBtu (thousand Btu)': 1194},
            'Lbs. (pounds)': {'kBtu (thousand Btu)': 1.194},
            'MLbs. (million pounds)': {'kBtu (thousand Btu)': 1194000},
            'GJ': {'kBtu (thousand Btu)': 947.817},
            'MBtu/ MMBtu/ Dth (million Btu/ dekatherm)': {'kBtu (thousand Btu)': 1000},
            'therms': {'kBtu (thousand Btu)': 100}
        },
        'District Hot Water': {
            'GJ': {'kBtu (thousand Btu)': 947.817},
            'MBtu/ MMBtu/ Dth (million Btu/ dekatherm)': {'kBtu (thousand Btu)': 1000},
            'therms': {'kBtu (thousand Btu)': 100}
        },
        'Electric - Grid': {
            'GJ': {'kBtu (thousand Btu)': 947.817},
            'MBtu/ MMBtu/ Dth (million Btu/ dekatherm)': {'kBtu (thousand Btu)': 1000},
            'kWh (thousand Watt-hours)': {'kBtu (thousand Btu)': 3.412},
            'MWh (million Watt-hours)': {'kBtu (thousand Btu)': 3412}
        },
        'Electric - Solar': {
            'GJ': {'kBtu (thousand Btu)': 947.817},
            'MBtu/ MMBtu/ Dth (million Btu/ dekatherm)': {'kBtu (thousand Btu)': 1000},
            'kWh (thousand Watt-hours)': {'kBtu (thousand Btu)': 3.412},
            'MWh (million Watt-hours)': {'kBtu (thousand Btu)': 3412}
        },
        'Electric - Wind': {
            'GJ': {'kBtu (thousand Btu)': 947.817},
            'MBtu/ MMBtu/ Dth (million Btu/ dekatherm)': {'kBtu (thousand Btu)': 1000},
            'kWh (thousand Watt-hours)': {'kBtu (thousand Btu)': 3.412},
            'MWh (million Watt-hours)': {'kBtu (thousand Btu)': 3412}
        },
        'District Chilled Water - Electric': {
            'GJ': {'kBtu (thousand Btu)': 947.817},
            'MBtu/ MMBtu/ Dth (million Btu/ dekatherm)': {'kBtu (thousand Btu)': 1000},
            'ton hours': {'kBtu (thousand Btu)': 12.0}
        },
        'District Chilled Water - Absorption': {
            'GJ': {'kBtu (thousand Btu)': 947.817},
            'MBtu/ MMBtu/ Dth (million Btu/ dekatherm)': {'kBtu (thousand Btu)': 1000},
            'ton hours': {'kBtu (thousand Btu)': 12.0}
        },
        'District Chilled Water - Engine': {
            'GJ': {'kBtu (thousand Btu)': 947.817},
            'MBtu/ MMBtu/ Dth (million Btu/ dekatherm)': {'kBtu (thousand Btu)': 1000},
            'ton hours': {'kBtu (thousand Btu)': 12.0}
        },
        'District Chilled Water - Other': {
            'GJ': {'kBtu (thousand Btu)': 947.817},
            'MBtu/ MMBtu/ Dth (million Btu/ dekatherm)': {'kBtu (thousand Btu)': 1000},
            'ton hours': {'kBtu (thousand Btu)': 12.0}
        }
    }
    return conversions[t][a][b]



#check to make sure the relevant sheets are in the file
def check_sheet(filepath, sheet):
    xl = pd.ExcelFile(filepath, engine='openpyxl')
    sheet_names = xl.sheet_names
    if sheet not in sheet_names:
        messages.error(f'Sheet {sheet} not found in Excel file.')
        st.stop()



#get the row to use for headers
def get_header_index(filepath, sheet):
    df = pd.read_excel(filepath, sheet, nrows=20)
    header_loc = df[df == 'Property Name'].dropna(axis=1, how='all').dropna(how='all')
    if not header_loc.shape[0] == 1:
        messages.error(f'Header row with Property Name column was not found.')
        st.stop()
    return header_loc.index.item()


#takes a list of buliding or meter objects and matches based on name or id
def match_obj(arr, id):
    if isinstance(arr[0], Building): 
        for building in arr:
            if building.name == id:
                return building
    elif isinstance(arr[0], Meter): 
        for meter in arr:
            if meter.id == id:
                return meter
    else:
        return None
    

#traverses the Excel file and create a list of Building classes
def get_buildings(filepath):
    
    check_sheet(filepath, 'Properties')
    check_sheet(filepath, 'Uses')

    buildings = []

    header_buildings = get_header_index(filepath, 'Properties') + 1
    df_buildings = pd.read_excel(filepath, sheet_name='Properties', header=header_buildings)

    #check to make sure the dataframe has all the necessary columns
    cols_to_check = [
        'Property Name', 
        'Street Address', 
        'Street Address 2', 
        'City/Municipality',
        'State/Province',
        'Postal Code',
        'Country',
        'Parent Property Name (if Applicable)'
    ]
    for c in cols_to_check:
        if c not in df_buildings.columns:
            messages.error(f'Missing column in Properties tab.')
            st.stop()

    
    
    for ind in df_buildings.index:
        building = Building(df_buildings['Property Name'][ind])
        
        address = df_buildings['Street Address'][ind]
        if df_buildings['Street Address 2'][ind] != 'Not Available':
            address += f', {df_buildings["Street Address 2"][ind]}'
        building.address = address
        
        building.city = df_buildings['City/Municipality'][ind]
        building.state = df_buildings['State/Province'][ind]
        building.postal = df_buildings['Postal Code'][ind]
        building.country = df_buildings['Country'][ind]
        
        if df_buildings['Parent Property Name (if Applicable)'][ind] != 'Not Available':
            building.parent = df_buildings['Parent Property Name (if Applicable)'][ind]
        
        buildings.append(building)
    
    header_uses = get_header_index(filepath, 'Uses') + 1
    df_uses = pd.read_excel(filepath, sheet_name='Uses', header=header_uses)

    #check to make sure the dataframe has all the necessary columns
    cols_to_check = [
        'Property Name', 
        'Use Type',
        'Gross Floor Area for Use',
        'Gross Floor Area Units'
    ]
    for c in cols_to_check:
        if c not in df_uses.columns:
            messages.error(f'Missing column in Uses tab.')
            st.stop()

    for ind in df_uses.index:
        name = df_uses['Property Name'][ind]
        building = match_obj(buildings, name)
        
        if building is not None:
            use_type = df_uses['Use Type'][ind]
            area = int(df_uses['Gross Floor Area for Use'][ind])
            area_units = df_uses['Gross Floor Area Units'][ind]
            if area_units != 'Sq. Ft.':
                try:
                    area *= convert('Area', area_units, 'Sq. Ft.')
                except:
                    building.add_note(f'Unable to convert area unit {area_units} to Sq. Ft.')
            
            building.add_use([use_type, area])
    
    return buildings


#traverses the Excel file and create a list of Meter classes
def get_meters(filepath):
    
    check_sheet(filepath, 'Meters')
    check_sheet(filepath, 'Meter Entries')

    meters = []

    header_meters = get_header_index(filepath, 'Meters') + 1
    df_meters = pd.read_excel(filepath, sheet_name='Meters', header=header_meters)

    #check to make sure the dataframe has all the necessary columns
    cols_to_check = [
        'Meter Type', 
        'Portfolio Manager Meter ID',
        'Property Name'
    ]
    for c in cols_to_check:
        if c not in df_meters.columns:
            messages.error(f'Missing column in Meters tab.')
            st.stop()
    

    for ind in df_meters.index:
        meter_type = df_meters['Meter Type'][ind]
        if Meter.valid_meter(meter_type):
            meter = Meter(df_meters['Portfolio Manager Meter ID'][ind])
            meter.building_name = df_meters['Property Name'][ind]
            meter.meter_type = meter_type
            meters.append(meter)
    
    header_entries = get_header_index(filepath, 'Meter Entries') + 1
    df_entries = pd.read_excel(filepath, sheet_name='Meter Entries', header=header_entries)

    #check to make sure the dataframe has all the necessary columns
    cols_to_check = [
        'Portfolio Manager Meter ID', 
        'Start Date',
        'End Date',
        'Delivery Date',
        'Meter Type',
        'Usage/Quantity',
        'Usage Units'
    ]
    for c in cols_to_check:
        if c not in df_entries.columns:
            messages.error(f'Missing column in Meter Entries tab.')
            st.stop()
            

    for ind in df_entries.index:
        id = df_entries['Portfolio Manager Meter ID'][ind]
        meter = match_obj(meters, id)
        if meter is not None:
            start = df_entries['Start Date'][ind]
            end = df_entries['End Date'][ind]
            delivery = df_entries['Delivery Date'][ind]
            fuel = df_entries['Meter Type'][ind]
            amount = df_entries['Usage/Quantity'][ind]
            units = df_entries['Usage Units'][ind]
        
            add_entry = True
            if units != 'kBtu (thousand Btu)':
                try:
                    amount *= convert(fuel, units, 'kBtu (thousand Btu)')
                except:
                    add_entry = False
                    meter.add_note(f'Unable to convert {fuel} with {units} to kBtu (thousand Btu); ')
            if add_entry:
                meter.add_entry(start=start, end=end, delivery=delivery, amount=amount)

    return meters


#take the meters in a buliding and convert the data to a simple monthly E, G, S, C, R frame
def resample_meters(building):
    meters = building.meters
    if len(meters) > 0:
        meter_type = meters[0].meter_type
        base = meters[0].entries.rename(columns={'kbtu': meter_type})
        if len(meters) > 1:
            for meter in meters[1:]:
                meter_type = meter.meter_type
                df = meter.entries.rename(columns={'kbtu': meter_type})
                base = base.add(df, fill_value=0).fillna(0)
        resampled = base.resample('M').sum()
        return resampled
    else:
        return None


#compile the monthly data and fill in missing data according to different methods
def compile_building_data(building, method='Blend'):

    entries = {}
    for u in ['E', 'G', 'S', 'C', 'R']:
        keys = list(map(lambda x: u + str(x), list(range(1, 13))))
        for k in keys:
            entries[k] = 0
    
    entries['year'] = None
    
    if len(building.meters) > 0:
        meter_types = [m.meter_type for m in building.meters]
        resampled = resample_meters(building)
        
        resampled['month'] = resampled.apply(lambda row: row.name.month, axis=1)
        resampled['year'] = resampled.apply(lambda row: row.name.year, axis=1)
        
        if method == 'Latest':
            #get data from the last calendar year
            year = datetime.date.today().year - 1
            df = resampled[resampled['year'] == year]
            index_by_month = df.set_index('month')
            for m in meter_types:
                series_dict = index_by_month[m].to_dict()
                for k in series_dict:
                    entries[f'{m}{k}'] = series_dict[k]
            entries['year'] = year
        
        elif method == 'Complete':
            #get data from the last *complete* calendar year
            meter_years = []
            for m in meter_types:
                #filter the dataframe to get only non-zero values
                non_zero = resampled[resampled[m] > 0]
                #group the values by year
                vals_by_year = list(non_zero.groupby('year')[m].count().to_dict().items())
                #sort the list by values ascending
                sorted_years = sorted(vals_by_year, key=lambda v: v[1])
                #grab the last item in the list, corresponding to the year with the most complete entries
                year = sorted_years[-1][0]
                meter_years.append(year)
                df = resampled[resampled['year'] == year]
                index_by_month = df.set_index('month')
                series_dict = index_by_month[m].to_dict()
                for k in series_dict:
                    entries[f'{m}{k}'] = series_dict[k]
            
            if len(meter_years) == 0:
                entries['year'] = ''
            elif len(meter_years) == 1:
                entries['year'] = meter_years[0]
            else:
                sorted_years = sorted(meter_years)
                entries['year'] = '-'.join([str(sorted_years[0]), str(sorted_years[-1])])
        
        else:
            meter_years = []
            #remove outliers, average data from last 5 years, fill in gaps, and smooth fuel deliveries
            for m in meter_types:
                #filter the dataframe to get only non-zero values
                non_zero = resampled[resampled[m] > 0]

                #see how many years there are with less than 4 months of data, which might indicate deliveries
                years = []
                grouped_year = non_zero.groupby('year')
                for year, group in grouped_year:
                    years.append([year, group[m].count()])
                
                number_of_years = len(years)
                years_with_possible_deliveries = len([y for y in years if y[1] <= 4])
                #if this is a trend, smooth out the values over the course of a year
                if years_with_possible_deliveries/number_of_years >= 0.5:
                    #get the sum of each year, divide by 12, and apply it to non_zero dataframe
                    for year in years:
                        total = non_zero.loc[non_zero['year'] == year[0], m].sum()
                        monthly = total/12
                        non_zero.loc[non_zero['year'] == year[0], m] = monthly

                #get a list of all values by month, sorted by year
                grouped_month = non_zero.groupby('month')
                for month, group in grouped_month:                    
                    #remove outliers
                    if len(group) > 1:
                        group_std = group[m].std()
                    else:
                        group_std = group[m]
                    
                    group_mean = group[m].mean()
                    filtered = group[group[m].apply(lambda row :abs(row-group_mean)<(group_std))]
                    #set index by year
                    index_by_year = filtered.set_index('year')
                    series_dict = index_by_year[m].to_dict()
                    sorted_by_year = sorted(list(series_dict.items()), key=lambda y: y[0])
                    #grab the last 5 years and take and average
                    latest_years = sorted_by_year[-5:]
                    meter_years = meter_years + [y[0] for y in latest_years]
                    mean_from_latest_years = sum([y[1] for y in latest_years])/len(latest_years)
                    entries[f'{m}{month}'] = mean_from_latest_years

                    if len(meter_years) == 0:
                        entries['year'] = ''
                    elif len(meter_years) == 1:
                        entries['year'] = meter_years[0]
                    else:
                        sorted_years = sorted(meter_years)
                        entries['year'] = '-'.join([str(sorted_years[0]), str(sorted_years[-1])])
            
    return entries


#create and save an excel file of records with standard header
def create_excel(records, local_name):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Inputs'
    #create headers
    headings = [
        ['Building Name', 'This is your unique building identifier.', 1],
        ['Area', 'Gross floor area.', 2],
        ['Building Type', 'The primary program type.', 3],
        ['Location', 'We use this information to locate the building.', 4],
        ['Year of Data', 'Energy data reference year.', 9],
        ['Heating System', 'Primary source of heating (optional)', 10],
        ['Cooling System', 'Primary source of cooling (optional)', 11],
        ['Electricity', 'The total amout of electricity consumed each month.', 12],
        ['Gas', 'The total amout of gas consumed each month.', 24],
        ['District Heating', 'The total amout of district steam or hot water consumed each month.', 36],
        ['District Cooling', 'The total amout of district chilled water consumed each month.', 48],
        ['PV Generation', 'The total amout of energy produced by onsite PV systems.', 60],
        ['PV Metering', 'How is the PV energy metered?', 72],
        ['Emissions Factors', "Optional - we'll use a default value based on location if you leave these blank.", 73],
        ['Leased', 'Is the building leased or owned?', 77],
        ['Notes', 'Any relevant notes about the building.', 78]
    ]
    units = [
        ['FT2', 2],
        ['kBtu', 12],
        ['kBtu', 24],
        ['kBtu', 36],
        ['kBtu', 48],
        ['kBtu', 60],
        ['Lbs CO2e / MBtu', 73],
        ['Lbs CO2e / MBtu', 74],
        ['Lbs CO2e / MBtu', 75],
        ['Lbs CO2e / MBtu', 76]
    ]
    subheadings = [
        ['ADDRESS', 4],
        ['CITY', 5],
        ['STATE', 6],
        ['COUNTRY', 7],
        *list(zip(['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC'], list(range(12, 24)))),
        *list(zip(['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC'], list(range(24, 36)))),
        *list(zip(['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC'], list(range(36, 48)))),
        *list(zip(['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC'], list(range(48, 60)))),
        *list(zip(['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC'], list(range(60, 72)))),
        ['ELECTRICITY', 73],
        ['GAS', 74],
        ['DISTRICT HEATING', 75],
        ['DISTRICT COOLING', 76]
    ]

    for heading in headings:
        col = get_column_letter(heading[2])
        sheet[col + '1'] = heading[0]
        sheet[col + '2'] = heading[1]
    
    for unit in units:
        col = get_column_letter(unit[1])
        sheet[col + '3'] = unit[0]
    
    for subheading in subheadings:
        col = get_column_letter(subheading[1])
        sheet[col + '4'] = subheading[0]    
    
    vals = []
    for u in ['E', 'G', 'S', 'C', 'R']:
        keys = list(map(lambda x: u + str(x), list(range(1, 13))))
        for k in keys:
            vals.append(k)

    for index, record in enumerate(records):
        row = index + 5
        sheet[get_column_letter(1) + str(row)] = record['building_name']
        sheet[get_column_letter(2) + str(row)] = record['area_ft2']
        sheet[get_column_letter(3) + str(row)] = record['buliding_type']
        sheet[get_column_letter(4) + str(row)] = record['address']
        sheet[get_column_letter(5) + str(row)] = record['city']
        sheet[get_column_letter(6) + str(row)] = record['state']
        sheet[get_column_letter(7) + str(row)] = record['country']
        sheet[get_column_letter(8) + str(row)] = record['zip']
        sheet[get_column_letter(9) + str(row)] = record['year']
        sheet[get_column_letter(10) + str(row)] = 'Unknown'
        sheet[get_column_letter(11) + str(row)] = 'Unknown'
        
        for index, val in enumerate(vals):
            col = index + 12
            sheet[get_column_letter(col) + str(row)] = record[val]
        
        sheet[get_column_letter(72) + str(row)] = 'Unknown'
        sheet[get_column_letter(73) + str(row)] = record['emissions_electricity']
        sheet[get_column_letter(74) + str(row)] = record['emissions_gas']
        sheet[get_column_letter(75) + str(row)] = record['emissions_district_heating']
        sheet[get_column_letter(76) + str(row)] = record['emissions_district_cooling']
        sheet[get_column_letter(77) + str(row)] = 'Unknown'
        sheet[get_column_letter(78) + str(row)] = record['notes']

    workbook.save(local_name)
    return local_name











#-------------------- MAIN CLASSES --------------------


#Holds all the buliding details
class Building:
    def __init__(self, name):
        self._name = name
        self._address = None
        self._city = None
        self._state = None
        self._postal = None
        self._country = None
        self._uses = []
        self._parent = None
        self._notes = None
        self._primary_use = None
        self._area_ft2 = 0
        self._meters = []
        self._fuel_types = {}
    
    @property
    def name(self):
        return self._name
    @name.setter
    def name(self, name):
        self._name = name
    
    @property
    def address(self):
        return self._address
    @address.setter
    def address(self, address):
        self._address = address
    
    @property
    def city(self):
        return self._city
    @city.setter
    def city(self, city):
        self._city = city

    @property
    def state(self):
        return self._state
    @state.setter
    def state(self, state):
        self._state = state

    @property
    def postal(self):
        return self._postal
    @postal.setter
    def postal(self, postal):
        self._postal = str(postal).zfill(5)

    @property
    def country(self):
        return self._country
    @country.setter
    def country(self, country):
        self._country = country
    
    @property
    def uses(self):
        return self._uses
    
    @property
    def primary_use(self):
        return self._primary_use
    
    @property
    def area_ft2(self):
        return self._area_ft2
    
    def add_use(self, use):
        self._uses.append(use)
        self._area_ft2 = sum([u[1] for u in self._uses])
        sorted_ascending_uses = sorted(self._uses, key=lambda u: u[1])
        self._primary_use = sorted_ascending_uses[-1][0]
            

    @property
    def parent(self):
        return self._parent
    @parent.setter
    def parent(self, parent):
        self._parent = parent
    
    @property
    def meters(self):
        return self._meters

    @property
    def fuel_types(self):
        return self._fuel_types
    
    def add_meter(self, meter):
        self._meters.append(meter)
        try:
            meter_type = meter.meter_type
            meter_fuel = meter.meter_fuel
            if meter_type in self._fuel_types:
                self._fuel_types[meter_type].append(meter_fuel)
            else:
                self._fuel_types[meter_type] = [meter_fuel]
        except:
            self.add_note(f'Problem parsing meter fuel and type.')
    
    @property
    def notes(self):
        return self._notes
    
    def add_note(self, note):
        if self._notes == None:
            self._notes = note
        else:
            self._notes = '; '.join([self._notes, note])



#Holds all the meter details. Gets attached to building.
class Meter:
    meters = {
        'Natural Gas': 'G',
        'Propane': 'G',
        'Fuel Oil (No. 2)': 'G',
        'Diesel': 'G',
        'District Steam': 'S',
        'District Hot Water': 'S',
        'Electric - Grid': 'E',
        'Electric - Solar': 'R',
        'Electric - Wind': 'R',
        'District Chilled Water - Electric': 'C',
        'District Chilled Water - Absorption': 'C',
        'District Chilled Water - Engine': 'C',
        'District Chilled Water - Other': 'C'
    }
    
    def __init__(self, id):
        self._id = id
        self._building_name = None
        self._meter_fuel = None
        self._meter_type = None
        self._entries = pd.DataFrame(columns=['kbtu'])
        self._notes = None

    @classmethod
    def valid_meter(cls, meter):
        return meter in Meter.meters
    
    @staticmethod
    def entry_validation(start, end, delivery, amount):
        validation = True
        message = ''
        try:
            if not (isinstance(start, datetime.datetime) or start == 'Not Available'):
                validation = False
                message += f'{start} is not a recognized; '
            if not (isinstance(end, datetime.datetime) or end == 'Not Available'):
                validation = False
                message += f'{end} is not a recognized; '
            if not (isinstance(delivery, datetime.datetime) or delivery == 'Not Available'):
                validation = False
                message += f'{delivery} is not a recognized; '
            if start == 'Not Available' and end == 'Not Available' and delivery == 'Not Available':
                validation = False
                message += 'No dates found; '
        except:
            validation = False
            message += 'Missing or invalid dates; '
        try: 
            float(amount)
        except:
            validation = False
            message += f'Missing or invalid amount; '
        return [validation, message]
    
    @property
    def id(self):
        return self._id
    @id.setter
    def id(self, id):
        self._id = id
    
    @property
    def building_name(self):
        return self._building_name
    @building_name.setter
    def building_name(self, name):
        self._building_name = name
    
    @property
    def meter_fuel(self):
        return self._meter_fuel
    
    @property
    def meter_type(self):
        return self._meter_type
    @meter_type.setter
    def meter_type(self, meter):
        if meter in list(Meter.meters.keys()):
            key = Meter.meters[meter]
            self._meter_type = key
            self._meter_fuel = meter
        else:
            messages.info(f'Meter of type {meter} is not recognized.')
    
    @property
    def notes(self):
        return self._notes
    
    def add_note(self, note):
        if self._notes == None:
            self._notes = note
        else:
            self._notes = '; '.join([self._notes, note])
    
    @property
    def entries(self):
        return self._entries
    
    def add_entry(self, start=None, end=None, delivery=None, amount=None):
 
        validation = self.entry_validation(start, end, delivery, amount)

        if not validation[0]:
            self.add_note(validation[1])
        else:
            if start == 'Not Available' and end == 'Not Available':
                dates = pd.date_range(delivery, delivery, freq='D')
            if start == 'Not Available' and delivery == 'Not Available':
                dates = pd.date_range(end, end, freq='D')
            if end == 'Not Available' and delivery == 'Not Available':
                dates = pd.date_range(start, start, freq='D')
            else:
                dates = pd.date_range(start, end, freq='D')
            
            zipped = zip(list(dates), [amount/len(list(dates))]*len(dates))
            df = pd.DataFrame(list(zipped), columns=['timestamp', 'kbtu'])
            reindexed = df.set_index('timestamp')
            
            self._entries = self._entries.add(reindexed, fill_value=0)








#-------------------- MAIN EXECUTION --------------------

def main(uploaded_file):
    buildings = get_buildings(uploaded_file)
    meters = get_meters(uploaded_file)
    for meter in meters:
        building_name = meter.building_name
        matched = match_obj(buildings, building_name)
        matched.add_meter(meter)
        matched.add_note(meter.notes)
    records = []
    for building in buildings:
        factors = get_factors(building.fuel_types, building.country, building.postal)
        building.add_note(factors['notes'])
        output = compile_building_data(building, method=fill_method)
        output['building_name'] = building.name
        output['area_ft2'] = building.area_ft2
        output['buliding_type'] = building.primary_use
        output['address'] = building.address
        output['city'] = building.city
        output['state'] = building.state
        output['country'] = building.country
        output['zip'] = building.postal
        output['notes'] = building.notes
        output['emissions_electricity'] = factors['E']
        output['emissions_gas'] = factors['G']
        output['emissions_district_heating'] = factors['S']
        output['emissions_district_cooling'] = factors['C']
        records.append(output)

    local_name = f'{uploaded_file.name.replace(".xlsx", "")}_export.xlsx'
    xl_path = create_excel(records, local_name)
    return xl_path



with form_container:
    st.subheader('Step 1')
    uploaded_file = st.file_uploader("Select .xslx output file.", type=['xlsx'], on_change=reset)

    st.subheader('Step 2')
    fill_method = st.selectbox('Choose how the metered energy data is compiled.', ('Blend', 'Complete', 'Latest'))

if uploaded_file is not None:
    messages.empty()
    button_container.empty()
    with st.spinner("Processing File..."):
        path = main(uploaded_file)
        with open(path, 'rb') as f:
            buffer = io.BytesIO(f.read())
    button_container.download_button('Download File', data=buffer, file_name=path, mime='application/vnd.ms-excel')
else:
    button_container.button('Download File', disabled=True)
